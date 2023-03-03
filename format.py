import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import json
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.styles import Alignment
import os

wrkbk_formatted = openpyxl.load_workbook("yelp_format_full2.xlsx")
wrkbk = openpyxl.load_workbook("yelp_data_full.xlsx")

greenFill = PatternFill(start_color='b6d7a8', end_color='b6d7a8', fill_type='solid')
pinkFill = PatternFill(start_color='ead1dc', end_color='ead1dc', fill_type='solid')
purpleFill = PatternFill(start_color='b4a7d6', end_color='b4a7d6', fill_type='solid')
blueFill = PatternFill(start_color='a4c2f4', end_color='a4c2f4', fill_type='solid')
lightgreenFill = PatternFill(start_color='d9ead3', end_color='d9ead3', fill_type='solid')
orangeFill = PatternFill(start_color='fce5cd', end_color='fce5cd', fill_type='solid')
redFill = PatternFill(start_color='e06666', end_color='e06666', fill_type='solid')
cyanFill = PatternFill(start_color='d0e0e3', end_color='d0e0e3', fill_type='solid')

for ws in list(wrkbk):
    city = ws.title
    ws_new = wrkbk_formatted.create_sheet(city)
    ws_new = wrkbk_formatted[city]
    
    ws_new.merge_cells(start_row=1, start_column=2, end_row=1, end_column=5)
    ws_new['A1'].fill = greenFill
    ws_new['A1'].alignment = Alignment(horizontal='center')
    ws_new['B1'].fill = greenFill
    ws_new['B1'].value = 'Activity'
    ws_new['B1'].alignment = Alignment(horizontal='center')
    ws_new['F1'].fill = greenFill
    ws_new['F1'].alignment = Alignment(horizontal='center')
    ws_new['A2'].fill = greenFill
    ws_new['A2'].value = 'Link from google'
    ws_new['A2'].alignment = Alignment(horizontal='center')
    ws_new['B2'].fill = greenFill
    ws_new['B2'].value = 'Activity Group'
    ws_new['B2'].alignment = Alignment(horizontal='center')
    ws_new['C2'].fill = greenFill
    ws_new['C2'].value = 'Activity Subgroup'
    ws_new['C2'].alignment = Alignment(horizontal='center')
    ws_new['D2'].fill = greenFill
    ws_new['D2'].value = 'Activity Type'
    ws_new['D2'].alignment = Alignment(horizontal='center')
    ws_new['E2'].fill = greenFill
    ws_new['E2'].value = 'Activity Name'
    ws_new['E2'].alignment = Alignment(horizontal='center')
    ws_new['F2'].fill = greenFill
    ws_new['F2'].value = 'Activity Description'
    ws_new['F2'].alignment = Alignment(horizontal='center')
    
    ws_new.merge_cells(start_row=1, start_column=7, end_row=1, end_column=10)
    ws_new['G1'].value = 'Age'
    ws_new['G1'].alignment = Alignment(horizontal='center')
    ws_new['G1'].fill = pinkFill
    ws_new['G2'].fill = pinkFill
    ws_new['G2'].value = 'Minimum age'
    ws_new['G2'].alignment = Alignment(horizontal='center')
    ws_new['H2'].fill = pinkFill
    ws_new['H2'].value = 'Max age'
    ws_new['H2'].alignment = Alignment(horizontal='center')
    ws_new['I2'].fill = pinkFill
    ws_new['I2'].value = 'Ideal Min age'
    ws_new['I2'].alignment = Alignment(horizontal='center')
    ws_new['J2'].fill = pinkFill
    ws_new['J2'].value = 'Ideal Max Age'
    ws_new['J2'].alignment = Alignment(horizontal='center')
    
    ws_new.merge_cells(start_row=1, start_column=11, end_row=1, end_column=12)
    ws_new['K1'].value = 'Price Per person'
    ws_new['K1'].alignment = Alignment(horizontal='center')
    ws_new['K1'].fill = purpleFill
    ws_new['K2'].fill = purpleFill
    ws_new['K2'].value = 'Minimum $'
    ws_new['K2'].alignment = Alignment(horizontal='center')
    ws_new['L2'].fill = purpleFill
    ws_new['L2'].value = 'Max $'
    ws_new['L2'].alignment = Alignment(horizontal='center')
    
    ws_new.merge_cells(start_row=1, start_column=13, end_row=1, end_column=17)
    ws_new['M1'].fill = blueFill
    ws_new['M1'].alignment = Alignment(horizontal='center')
    ws_new['M1'].value = 'Business Details'
    ws_new['M2'].fill = blueFill
    ws_new['M2'].value = 'Business Name'
    ws_new['M2'].alignment = Alignment(horizontal='center')
    ws_new['N2'].fill = blueFill
    ws_new['N2'].value = 'Website'
    ws_new['N2'].alignment = Alignment(horizontal='center')
    ws_new['O2'].fill = blueFill
    ws_new['O2'].value = 'Address'
    ws_new['O2'].alignment = Alignment(horizontal='center')
    ws_new['P2'].fill = blueFill
    ws_new['P2'].value = 'City'
    ws_new['P2'].alignment = Alignment(horizontal='center')
    ws_new['Q2'].fill = blueFill
    ws_new['Q2'].value = 'Link to Photos for cover image etc'
    ws_new['Q2'].alignment = Alignment(horizontal='center')
    
    ws_new.merge_cells(start_row=1, start_column=18, end_row=1, end_column=20)
    ws_new['R1'].fill = lightgreenFill
    ws_new['R1'].alignment = Alignment(horizontal='center')
    ws_new['R1'].value = 'Business Contact'
    ws_new['R2'].fill = lightgreenFill
    ws_new['R2'].value = 'Phone Number'
    ws_new['R2'].alignment = Alignment(horizontal='center')
    ws_new['S2'].fill = lightgreenFill
    ws_new['S2'].value = 'Email'
    ws_new['S2'].alignment = Alignment(horizontal='center')
    ws_new['T2'].fill = lightgreenFill
    ws_new['T2'].value = 'Contact Us Page'
    ws_new['T2'].alignment = Alignment(horizontal='center')  

    ws_new.merge_cells(start_row=1, start_column=21, end_row=1, end_column=37)
    ws_new['U1'].fill = orangeFill
    ws_new['U1'].alignment = Alignment(horizontal='center')
    ws_new['U1'].value = 'Activity Details'
    ws_new['U2'].fill = orangeFill
    ws_new['U2'].value = 'Activity Location Type'
    ws_new['U2'].alignment = Alignment(horizontal='center')
    ws_new['V2'].fill = orangeFill
    ws_new['V2'].value = 'Activity Duration'
    ws_new['V2'].alignment = Alignment(horizontal='center')
    ws_new['W2'].fill = orangeFill
    ws_new['W2'].value = 'Activity Start Time'
    ws_new['W2'].alignment = Alignment(horizontal='center')
    ws_new['X2'].fill = orangeFill
    ws_new['X2'].value = 'Activity End Time'
    ws_new['X2'].alignment = Alignment(horizontal='center') 
    ws_new.merge_cells(start_row=2, start_column=25, end_row=2, end_column=31)  
    ws_new['Y2'].fill = orangeFill
    ws_new['Y2'].value = 'Activity Days'
    ws_new['Y2'].alignment = Alignment(horizontal='center')    
    
    ws_new['Y3'].value = "Monday Hours"
    ws_new['Y3'].alignment = Alignment(horizontal='center')  
    ws_new['Z3'].value = "Tuesday Hours"
    ws_new['Z3'].alignment = Alignment(horizontal='center')  
    ws_new['AA3'].value = "Wednesday Hours"
    ws_new['AA3'].alignment = Alignment(horizontal='center')  
    ws_new['AB3'].value = "Thursday Hours"
    ws_new['AB3'].alignment = Alignment(horizontal='center')  
    ws_new['AC3'].value = "Friday Hours"
    ws_new['AC3'].alignment = Alignment(horizontal='center')  
    ws_new['AD3'].value = "Saturday Hours"
    ws_new['AD3'].alignment = Alignment(horizontal='center') 
    ws_new['AE3'].value = "Sunday Hours"
    ws_new['AE3'].alignment = Alignment(horizontal='center') 
    
    ws_new['AF2'].fill = orangeFill
    ws_new['AF2'].value = 'Activity Start Date'
    ws_new['AF2'].alignment = Alignment(horizontal='center')        
    ws_new['AG2'].fill = orangeFill
    ws_new['AG2'].value = 'Activity End Date'
    ws_new['AG2'].alignment = Alignment(horizontal='center')     
    ws_new['AH2'].fill = orangeFill
    ws_new['AH2'].value = 'Need Reservation?'
    ws_new['AH2'].alignment = Alignment(horizontal='center') 
    ws_new['AI2'].fill = orangeFill
    ws_new['AI2'].value = 'Availability (yes/no for classes)'
    ws_new['AI2'].alignment = Alignment(horizontal='center')   
    ws_new['AJ2'].fill = orangeFill
    ws_new['AJ2'].value = 'Availability/Ticketing/Reservation link'
    ws_new['AJ2'].alignment = Alignment(horizontal='center')   
    ws_new['AK2'].fill = orangeFill
    ws_new['AK2'].value = 'Google Review Link'
    ws_new['AK2'].alignment = Alignment(horizontal='center')    

    ws_new.merge_cells(start_row=1, start_column=38, end_row=1, end_column=39)
    ws_new['AL1'].fill = redFill
    ws_new['AL1'].alignment = Alignment(horizontal='center')
    ws_new['AL1'].value = 'membership'
    ws_new['AL2'].fill = redFill
    ws_new['AL2'].value = 'Do they offer membership?'
    ws_new['AL2'].alignment = Alignment(horizontal='center')
    ws_new['AM2'].fill = redFill
    ws_new['AM2'].value = 'Membership $ per person'
    ws_new['AM2'].alignment = Alignment(horizontal='center')  

    ws_new.merge_cells(start_row=1, start_column=40, end_row=1, end_column=43)
    ws_new['AN1'].fill = cyanFill
    ws_new['AN1'].alignment = Alignment(horizontal='center')
    ws_new['AN1'].value = 'Social Media Links'
    ws_new['AN2'].fill = cyanFill
    ws_new['AN2'].value = 'Yelp'
    ws_new['AN2'].alignment = Alignment(horizontal='center')
    ws_new['AO2'].fill = cyanFill
    ws_new['AO2'].value = 'Facebook'
    ws_new['AO2'].alignment = Alignment(horizontal='center')
    ws_new['AP2'].fill = cyanFill
    ws_new['AP2'].value = 'Instagram'
    ws_new['AP2'].alignment = Alignment(horizontal='center')   
    ws_new['AQ2'].fill = cyanFill
    ws_new['AQ2'].value = 'Others'
    ws_new['AQ2'].alignment = Alignment(horizontal='center') 


    i = 1
    row_new = 3
    days_dict = {}
    for line in ws.iter_rows():
                
        if i != 1:
            #link
            ws_new.cell(row=row_new, column=1).value = ws.cell(row=i, column=1).value
            #search term
            ws_new.cell(row=row_new, column=2).value = ws.cell(row=i, column=2).value
            #name
            ws_new.cell(row=row_new, column=5).value = ws.cell(row=i, column=4).value
            #
            #ws_new.cell(row=row_new, column=13).value = ws.cell(row=i, column=2).value
            #website
            ws_new.cell(row=row_new, column=14).value = ws.cell(row=i, column=14).value
            #address
            ws_new.cell(row=row_new, column=15).value = ws.cell(row=i, column=5).value
            #city
            ws_new.cell(row=row_new, column=16).value = city
            #phone
            ws_new.cell(row=row_new, column=18).value = ws.cell(row=i, column=15).value
            #reservation
            ws_new.cell(row=row_new, column=34).value = ws.cell(row=i, column=13).value
            
            ws_new.cell(row=row_new, column=25).value = ws.cell(row=i, column=6).value
            ws_new.cell(row=row_new, column=26).value = ws.cell(row=i, column=7).value
            ws_new.cell(row=row_new, column=27).value = ws.cell(row=i, column=8).value
            ws_new.cell(row=row_new, column=28).value = ws.cell(row=i, column=9).value
            ws_new.cell(row=row_new, column=29).value = ws.cell(row=i, column=10).value
            ws_new.cell(row=row_new, column=30).value = ws.cell(row=i, column=11).value
            ws_new.cell(row=row_new, column=31).value = ws.cell(row=i, column=12).value
            
            #category
            ws_new.cell(row=row_new, column=3).value = ws.cell(row=i, column=3).value
            #review link
            #ws_new.cell(row=row_new, column=37).value = ws.cell(row=i, column=15).value            
            
            
        i += 1
        row_new += 1
            
    
wrkbk_formatted.save("yelp_format_full2.xlsx")