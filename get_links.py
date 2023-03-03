import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import json
import openpyxl
import os

#WEBDRIVER
driver = webdriver.Chrome('chromedriver.exe')

cities = []
#GET CITIES
with open("cities.txt", "r") as f:
    for line in f:
        cities.append(line.strip())
   
categories =  []
with open("categories.txt", "r") as f:
    for line in f:
        categories.append(line.strip())


wrkbk = openpyxl.load_workbook("yelp_data_full.xlsx")

for city in cities:
    myArr = city.split('::')
    city = myArr[0]
    coordinates = myArr[1]
    print(city)
    myDict = {}
    ws = wrkbk.create_sheet(city)
    ws = wrkbk[city]
    
    ws.cell(row=1, column=1).value = "Link"
    ws.cell(row=1, column=2).value = "Category"
    ws.cell(row=1, column=3).value = "Tags"
    
    row_num = 2
    for category in categories:

        myArr = category.split('::')
        category = myArr[0]
        tag_to_find = myArr[1]
        category_to_link = "+".join(category.strip().split(" "))
        print(category_to_link)
        city_to_link = "+".join(city.strip().split(" "))
        print(city_to_link)
        link = "https://www.yelp.com/search?find_desc="+category_to_link+"&find_loc="+city_to_link+"%2C+CA%2C+United+States" + coordinates
        driver.get(link)
        
        time.sleep(2)
        
        
        pages = driver.find_elements(By.CLASS_NAME, "border-color--default__09f24__NPAKY.text-align--center__09f24__fYBGO")
        for no in pages:
            if "1 of"  in no.get_attribute("textContent"):
                no_of_pages = no.get_attribute("textContent").replace("1 of ", "")
                print(no_of_pages)
                if no_of_pages != '':
                
                    for page in range(min(int(no_of_pages), 8)):
                        
                        attr = str(page*10)
                        
                        driver.get(link + "&start=" + attr)
                        cards = driver.find_elements(By.CLASS_NAME, "container__09f24__mpR8_")
                        print(len(cards))
                        for card in cards:
                            print(card)
                            name = card.find_element(By.CLASS_NAME, "css-1agk4wl")
                            url = name.find_element(By.XPATH, "./span/a").get_attribute("href")
                            url_to_append = url.split("osq")[0]
                            tags_big = card.find_element(By.CLASS_NAME, "css-dzq7l1").find_elements(By.CLASS_NAME, "display--inline__09f24__c6N_k")
                            if len(tags_big) != 0:
                                tags_big[0].find_elements(By.TAG_NAME, "a")
                            tags = []
                            for tag in tags_big:
                                tags.append(tag.get_attribute("textContent"))
                            tags= ",".join(tags)
                            print(tags)
                            if 'osq' in url and url_to_append not in myDict:
                                if tag_to_find == '' or tag_to_find in tags:
                                    ws.cell(row=row_num, column=1).value = url
                                    ws.cell(row=row_num, column=2).value = category
                                    ws.cell(row=row_num, column=3).value = tags
                                    myDict[url_to_append] = row_num                                               
                                    row_num += 1
                            elif url_to_append in myDict: 
                                cat = ws.cell(row=myDict[url_to_append], column=2).value
                                print(cat)
                                cat = cat + "," + category 
                                ws.cell(row=myDict[url_to_append], column=2).value = cat
                                
        wrkbk.save("yelp_data2.xlsx")
            

   
        
driver.close()
wrkbk.save("yelp_data2.xlsx")