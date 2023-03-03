import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import json
import openpyxl
import os

wrkbk = openpyxl.load_workbook("yelp_data_full.xlsx")
driver = webdriver.Chrome('chromedriver.exe')


try:
    for ws in list(wrkbk):

        i = 1
        ws.cell(row=i, column=4).value = 'Name'
        ws.cell(row=i, column=5).value = 'Address'
        ws.cell(row=i, column=6).value = 'Monday hours'
        ws.cell(row=i, column=7).value = 'Tuesday hours'
        ws.cell(row=i, column=8).value = 'Wednesday hours'
        ws.cell(row=i, column=9).value = 'Thursday hours'
        ws.cell(row=i, column=10).value = 'Friday hours'
        ws.cell(row=i, column=11).value = 'Saturday hours'
        ws.cell(row=i, column=12).value = 'Sunday hours'
        ws.cell(row=i, column=13).value = 'Need Reservation?'
        ws.cell(row=i, column=14).value = 'Website'
        ws.cell(row=i, column=15).value = 'Phone'
        for row in ws.iter_rows():
            #print(row[13].value)
            if i != 1 and (row[4].value == '' or row[4].value == None):
                link = row[0].value
                print(ws.title)
                print(i)
                driver.get(link)
                #time.sleep(2)
                try:
                    name = driver.find_element(By.CLASS_NAME, "css-1se8maq").get_attribute("textContent").strip()     
                    ws.cell(row=i, column=4).value = name
                    address = driver.find_element(By.TAG_NAME, "address").get_attribute("textContent").strip()
                    ws.cell(row=i, column=5).value = address
                    hours = driver.find_elements(By.CLASS_NAME, "no-wrap__09f24__c3plq.css-1p9ibgf")
                    for h in range(7):
                        ws.cell(row=i, column=(6+h)).value = hours[h].get_attribute("textContent")
                    
                except:
                    pass
                    
                    
                try:
                    amenities = driver.find_element(By.CLASS_NAME, "layout-wrap__09f24__GEBlv.layout-2-units__09f24__PsGVW")
                    amenities_text = amenities.get_attribute("textContent")
                    
                    if "By Appointment Only" in amenities_text:
                        ws.cell(row=i, column=13).value = "yes"
                
                except Exception as e:
                    pass
                
                try:    
                    contacts = []
                    try:
                        contacts =  driver.find_element(By.XPATH, "/html/body/yelp-react-root/div[1]/div[4]/div/div/div[2]/div/div[2]/div").find_elements(By.CLASS_NAME, "css-1p9ibgf")
                    except:
                        try:
                            contacts =  driver.find_element(By.XPATH, "/html/body/yelp-react-root/div[1]/div[3]/div/div/div[2]/div/div[2]/div").find_elements(By.CLASS_NAME, "css-1p9ibgf")
                        except:
                            pass
                    telephone = contacts[0].get_attribute("textContent")
                    print(contacts[1].get_attribute("textContent"))

                    website = contacts[0].find_element(By.XPATH, "./a").get_attribute("href")

                    if "…" in contacts[0].get_attribute("textContent"):
                        driver.get(website)
                        time.sleep(3)
                        print(driver.current_url)
                        ws.cell(row=i, column=14).value = driver.current_url
                    ws.cell(row=i, column=15).value = telephone
                    
                    
                except Exception as e:
                    print(e)
                    pass
            i+= 1

            wrkbk.save("yelp_data_full.xlsx")
            time.sleep(2)    
except Exception as e:     
    print(e)
    wrkbk.save("yelp_data_full.xlsx")
    
wrkbk.save("yelp_data_full.xlsx")

